"""Excel export using openpyxl."""
from io import BytesIO
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from attendance.models import DailySummary, LatenessRecord, AttendanceLog
from penalties.models import Penalty


def export_attendance_excel(start: date, end: date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = ["Date", "Employee ID", "Name", "Status", "Check In", "Check Out", "Working (min)", "Minutes Late", "Missing Check Out"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)
    rows = DailySummary.objects.filter(date__gte=start, date__lte=end).select_related("employee").order_by("date", "employee__employee_id")
    for row_idx, s in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=str(s.date))
        ws.cell(row=row_idx, column=2, value=s.employee.employee_id)
        ws.cell(row=row_idx, column=3, value=s.employee.get_full_name())
        ws.cell(row=row_idx, column=4, value=s.get_status_display())
        ws.cell(row=row_idx, column=5, value=s.check_in_time.strftime("%H:%M") if s.check_in_time else "")
        ws.cell(row=row_idx, column=6, value=s.check_out_time.strftime("%H:%M") if s.check_out_time else "")
        ws.cell(row=row_idx, column=7, value=s.working_minutes)
        ws.cell(row=row_idx, column=8, value=s.minutes_late)
        ws.cell(row=row_idx, column=9, value="Yes" if s.missing_check_out else "No")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_lateness_excel(start: date, end: date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lateness"
    headers = ["Date", "Employee ID", "Name", "Minutes Late", "Check In Time", "Expected Start"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)
    rows = LatenessRecord.objects.filter(date__gte=start, date__lte=end).select_related("employee").order_by("date", "employee__employee_id")
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=str(r.date))
        ws.cell(row=row_idx, column=2, value=r.employee.employee_id)
        ws.cell(row=row_idx, column=3, value=r.employee.get_full_name())
        ws.cell(row=row_idx, column=4, value=r.minutes_late)
        ws.cell(row=row_idx, column=5, value=r.check_in_time.strftime("%Y-%m-%d %H:%M") if r.check_in_time else "")
        ws.cell(row=row_idx, column=6, value=str(r.expected_start))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_penalty_excel(start: date, end: date, employee_id: Optional[str] = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Penalties"
    headers = ["Date", "Employee ID", "Name", "Amount", "Percent", "Rule", "Reason", "Manual"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)
    rows = Penalty.objects.filter(penalty_date__gte=start, penalty_date__lte=end).select_related("employee", "rule")
    if employee_id:
        rows = rows.filter(employee__employee_id=employee_id)
    rows = rows.order_by("penalty_date", "created_at")
    for row_idx, p in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=p.penalty_date.strftime("%Y-%m-%d") if p.penalty_date else "")
        ws.cell(row=row_idx, column=2, value=p.employee.employee_id)
        ws.cell(row=row_idx, column=3, value=p.employee.get_full_name())
        ws.cell(row=row_idx, column=4, value=float(p.amount))
        ws.cell(row=row_idx, column=5, value=float(p.penalty_percent) if p.penalty_percent is not None else "")
        ws.cell(row=row_idx, column=6, value=p.rule.name if p.rule else "")
        ws.cell(row=row_idx, column=7, value=p.reason or "")
        ws.cell(row=row_idx, column=8, value="Yes" if p.is_manual else "No")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_attendance_logs_excel(
    start: date,
    end: date,
    employee_id: Optional[str] = None,
    event_type: Optional[str] = None,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    headers = ["DateTime", "Employee ID", "Name", "Event", "Source", "Source ID"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)
    qs = AttendanceLog.objects.filter(timestamp__date__gte=start, timestamp__date__lte=end).select_related("employee")
    if employee_id:
        qs = qs.filter(employee__employee_id=employee_id)
    if event_type:
        qs = qs.filter(event_type=event_type)
    qs = qs.order_by("-timestamp")
    for row_idx, log in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1, value=log.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_idx, column=2, value=log.employee.employee_id)
        ws.cell(row=row_idx, column=3, value=log.employee.get_full_name())
        ws.cell(row=row_idx, column=4, value=log.get_event_type_display())
        ws.cell(row=row_idx, column=5, value=log.source or "")
        ws.cell(row=row_idx, column=6, value=log.source_id or "")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
